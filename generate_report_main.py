import os
import json
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

ALLURE_RESULTS_DIR = "./report/tmp"


def load_attachment(source: str) -> str:
    """读取 allure 的附件内容（支持 JSON 和文本）"""
    filepath = os.path.join(ALLURE_RESULTS_DIR, source)
    if not os.path.exists(filepath):
        return "-"

    with open(filepath, "r", encoding="utf-8") as f:
        try:
            return json.dumps(json.load(f), ensure_ascii=False, indent=2)
        except Exception:
            f.seek(0)
            return f.read()


def parse_allure_results(results_dir: str):
    """解析 allure 结果目录，返回用例明细和请求信息"""
    total, passed, failed, skipped, duration = 0, 0, 0, 0, 0
    case_details, api_steps = [], []

    for file in os.listdir(results_dir):
        if not file.endswith("result.json"):
            continue

        with open(os.path.join(results_dir, file), "r", encoding="utf-8") as f:
            data = json.load(f)

        # 统计信息
        total += 1
        status = data.get("status", "-")
        start, stop = data.get("start", 0), data.get("stop", 0)
        duration += stop - start

        if status == "passed":
            passed += 1
        elif status == "failed":
            failed += 1
        elif status == "skipped":
            skipped += 1

        # 错误信息
        status_details = data.get("statusDetails", {})
        error_message = status_details.get("message", "-")
        error_trace = status_details.get("trace", "-")

        # 标签（epic, feature, story）
        labels = {l["name"]: l["value"] for l in data.get("labels", [])}
        epic = labels.get("epic", "-")
        feature = labels.get("feature", "-")
        story = labels.get("story", "-")
        method = data.get("fullName", "").split("#")[-1]

        # 参数
        parameters = data.get("parameters", [])
        param_str = ", ".join([f"{p['name']}={p['value']}" for p in parameters]) or "-"

        # 用例明细
        case_details.append(
            [
                epic,
                feature,
                story,
                method,
                data.get("name", "-"),
                param_str,
                status,
                stop - start,
                error_message,
                error_trace,
            ]
        )

        # 请求步骤
        url, method_type, headers, req_data, exp_data, resp_time, resp_data = (
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
        )
        for step in data.get("steps", []):
            step_name = step["name"]
            if "URL" in step_name:
                url = step_name.split("URL: ")[-1]
            elif "请求方式" in step_name:
                method_type = step_name.split(": ")[-1]
            elif "请求头" in step_name:
                headers = (
                    load_attachment(step["attachments"][0]["source"])
                    if step.get("attachments")
                    else "-"
                )
            elif "请求数据" in step_name:
                req_data = (
                    load_attachment(step["attachments"][0]["source"])
                    if step.get("attachments")
                    else "-"
                )
            elif "预期数据" in step_name:
                exp_data = (
                    load_attachment(step["attachments"][0]["source"])
                    if step.get("attachments")
                    else "-"
                )
            elif "响应耗时" in step_name:
                resp_time = step_name.split(": ")[-1]
            elif "响应结果" in step_name:
                resp_data = (
                    load_attachment(step["attachments"][0]["source"])
                    if step.get("attachments")
                    else "-"
                )

        api_steps.append(
            [
                data.get("name", "-"),
                url,
                method_type,
                headers,
                req_data,
                exp_data,
                resp_time,
                resp_data,
            ]
        )

    return {
        "stats": {
            "total": total,
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
            "duration": duration,
        },
        "case_details": case_details,
        "api_steps": api_steps,
    }


def generate_api_report(
    results_dir=ALLURE_RESULTS_DIR, output_file="接口测试报告.xlsx"
):
    """生成接口测试 Excel 报告"""
    results = parse_allure_results(results_dir)
    wb = openpyxl.Workbook()

    # --- Sheet1：测试概览 ---
    ws_summary = wb.active
    ws_summary.title = "测试概览"
    ws_summary.append(["指标", "数量"])
    stats = results["stats"]
    ws_summary.append(["用例总数", stats["total"]])
    ws_summary.append(["通过", stats["passed"]])
    ws_summary.append(["失败", stats["failed"]])
    ws_summary.append(["跳过", stats["skipped"]])
    ws_summary.append(
        [
            "通过率",
            f"{(stats['passed']/stats['total']*100):.2f}%" if stats["total"] else "0%",
        ]
    )
    ws_summary.append(["执行时长(秒)", round(stats["duration"] / 1000, 2)])
    ws_summary.append(["测试时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    # --- Sheet2：用例明细 ---
    ws_detail = wb.create_sheet("用例明细")
    ws_detail.append(
        [
            "Epic",
            "Feature",
            "Story",
            "方法",
            "用例标题",
            "参数",
            "状态",
            "耗时(ms)",
            "错误信息",
            "异常堆栈",
        ]
    )
    status_colors = {
        "passed": PatternFill("solid", fgColor="90EE90"),  # 浅绿色
        "failed": PatternFill("solid", fgColor="FF7F7F"),  # 浅红色
        "skipped": PatternFill("solid", fgColor="FFD966"),  # 黄色
        "-": PatternFill("solid", fgColor="D9D9D9"),  # 灰色
    }

    for row in sorted(results["case_details"], key=lambda x: (x[0], x[3])):
        ws_detail.append(row)
        status_cell = ws_detail.cell(row=ws_detail.max_row, column=7)
        status = status_cell.value
        if status in status_colors:
            status_cell.fill = status_colors[status]

    # --- Sheet3：接口请求信息 ---
    ws_api = wb.create_sheet("接口请求信息")
    ws_api.append(
        [
            "用例标题",
            "请求URL",
            "请求方式",
            "请求头",
            "请求数据",
            "预期数据",
            "响应耗时",
            "响应结果",
        ]
    )
    for row in sorted(results["api_steps"], key=lambda x: x[0]):
        ws_api.append(row)

    # 自动列宽
    for sheet in [ws_summary, ws_detail, ws_api]:
        for col in sheet.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value), default=10
            )
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    wb.save(output_file)
    print(f"✅ 接口测试报告已生成: {output_file}")


if __name__ == "__main__":
    generate_api_report()
